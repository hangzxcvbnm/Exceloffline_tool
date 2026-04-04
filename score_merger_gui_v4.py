#!/usr/bin/env python3
"""
Excel 多文件合并工具（GUI v4.0 多文件合并版）
- 支持动态添加/删除多个 Excel 文件
- 支持"纵向追加"（stack）和"横向合并"（join）两种模式
- 合并键可选（默认学号/工号）
- 离线运行，无需联网
- 自动识别常见文件名
- 全新视觉排版（v3 精致卡片风格保留）
- 适合作为 Windows EXE 直接分发

- 打包 Windows EXE:
  py -m PyInstaller --noconfirm --clean --onefile --windowed --icon=score_merger_icon.ico --name 多文件合并工具 score_merger_gui_v4.py
"""

import os
import sys
import ctypes
import traceback
import threading
from pathlib import Path
from dataclasses import dataclass
from typing import List, Optional
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

APP_TITLE = "Excel 多文件合并工具"
APP_VERSION = "v4.0 多文件版"
APP_SUBTITLE = "离线多表合并桌面工具"
DEFAULT_FOLDER = Path.home() / "Desktop" / "表格输入输出" / "表格导入导出"
OUTPUT_NAME = "合并结果.xlsx"
UI_SCALE = 1.0

# ============================================================
#  配色系统 - 毛玻璃现代风格 (Glassmorphism)
# ============================================================
C = {
    # 背景 - 柔和的渐变灰
    "bg":            "#F0F4F8",
    "bg_gradient":   "#E8EEF5",
    
    # 卡片 - 半透明白色
    "card":          "#FFFFFF",
    "card_alpha":    "#FFFFFF",  # 会配合透明度使用
    "card_hover":    "#F8FAFC",
    
    # 毛玻璃效果色
    "glass_bg":      "rgba(255, 255, 255, 0.72)",
    "glass_border":  "rgba(255, 255, 255, 0.5)",
    "glass_shadow":  "rgba(31, 45, 61, 0.08)",
    
    # 主色调 - 清新的科技蓝
    "banner":        "#3B82F6",
    "banner_light":  "#60A5FA",
    "banner_soft":   "#DBEAFE",
    
    # 强调色
    "accent":        "#6366F1",  # 靛蓝
    "accent_light":  "#E0E7FF",
    
    # 文字
    "text_primary":  "#1E293B",
    "text_secondary":"#64748B",
    "text_muted":    "#94A3B8",
    "text_inverse":  "#FFFFFF",
    
    # 功能色 - 更柔和的版本
    "blue":          "#3B82F6",
    "blue_light":    "#DBEAFE",
    "blue_soft":     "#EFF6FF",
    
    "green":         "#10B981",
    "green_text":    "#059669",
    "green_soft":    "#ECFDF5",
    
    "orange":        "#F59E0B",
    "orange_text":   "#D97706",
    "orange_soft":   "#FFFBEB",
    
    "red":           "#EF4444",
    "red_text":      "#DC2626",
    "red_soft":      "#FEF2F2",
    
    "purple":        "#8B5CF6",
    
    # 边框和分割线
    "border":        "#E2E8F0",
    "border_light":  "#F1F5F9",
    "blue_border":   "#BFDBFE",
    
    # 终端/日志区域
    "term_bg":       "#0F172A",
    
    # 其他兼容色
    "action_bg":     "#EFF6FF",
    "card_alt":      "#F8FAFC",
    "banner_deep":   "#1E40AF",
}

# ============================================================
#  FileItem
# ============================================================
@dataclass
class FileItem:
    path: str
    name: str
    status: str = "已添加"


# ============================================================
#  字体 / DPI / 高 DPI 自适应
# ============================================================
def ui_font(size, weight="normal", family="Microsoft YaHei UI"):
    scaled = max(8, int(round(size * UI_SCALE)))
    return (family, scaled, weight)

def mono_font(size=9):
    scaled = max(9, int(round(size * UI_SCALE)))
    return ("Cascadia Mono", scaled)

def enable_high_dpi():
    if not sys.platform.startswith("win"):
        return
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass

def detect_ui_scale():
    if not sys.platform.startswith("win"):
        return 1.04
    try:
        user32 = ctypes.windll.user32
        gdi32 = ctypes.windll.gdi32
        hwnd = user32.GetDesktopWindow()
        dc = user32.GetDC(hwnd)
        try:
            dpi_x = gdi32.GetDeviceCaps(dc, 88)
        finally:
            user32.ReleaseDC(hwnd, dc)
        scale = dpi_x / 96.0
        scale = max(1.0, min(scale * 1.04, 1.5))
        return round(scale, 2)
    except Exception:
        return 1.2

def get_icon_path() -> Optional[str]:
    if getattr(sys, "frozen", False):
        base = Path(sys._MEIPASS)
    else:
        base = Path(__file__).parent
    for name in ["score_merger_icon.ico", "icon.ico"]:
        p = base / name
        if p.exists():
            return str(p)
    return None

def set_window_icon(root: tk.Tk):
    icon = get_icon_path()
    if not icon or not sys.platform.startswith("win"):
        return
    try:
        root.iconbitmap(icon)
    except Exception:
        pass
    try:
        img = tk.PhotoImage(file=icon)
        root.iconphoto(True, img)
    except Exception:
        pass

def configure_window(root: tk.Tk):
    try:
        root.update_idletasks()
        sw = root.winfo_screenwidth()
        sh = root.winfo_screenheight()
        tw = min(1180, max(980, sw - 120))
        th = min(820, max(700, sh - 120))
        mw = min(960, tw)
        mh = min(680, th)
        px = max((sw - tw) // 2, 20)
        py = max((sh - th) // 2, 20)
        root.geometry(f"{tw}x{th}+{px}+{py}")
        root.minsize(mw, mh)
    except Exception:
        root.geometry("1120x780")
        root.minsize(960, 680)


# ============================================================
#  主应用
# ============================================================
class MultiMergeGUI:
    def __init__(self, root):
        global UI_SCALE
        UI_SCALE = detect_ui_scale()

        self.root = root
        self.root.title(f" {APP_TITLE} {APP_VERSION}")
        set_window_icon(self.root)
        configure_window(self.root)
        self.root.configure(bg=C["bg"])

        # 状态
        self.file_items: List[FileItem] = []
        self.output_path = tk.StringVar()
        self.merge_mode = tk.StringVar(value="stack")     # stack=纵向, join=横向
        self.join_key = tk.StringVar(value="学号/工号")   # join 模式的列名
        self.column_filter = tk.StringVar(value="")       # 横向合并时指定要合并的列（逗号分隔）
        self.status_text = tk.StringVar(value="就绪")
        self.summary_text = tk.StringVar(value="添加文件后即可开始")

        # UI 元素引用
        self.file_list_frame = None
        self.file_count_var = tk.StringVar(value="0 个文件")
        self.log_text = None
        self.merge_btn = None

        self._apply_style()
        self._build_ui()
        self._auto_scan_defaults()

    # ────────── 样式 ──────────
    def _apply_style(self):
        style = ttk.Style()
        try:
            style.theme_use("vista")
        except Exception:
            pass
        style.configure("Primary.TButton",
                        font=ui_font(11, "bold"), padding=(24, 10), foreground=C["blue"])
        style.map("Primary.TButton",
                  background=[("active", C["blue_light"])])
        style.configure("Ghost.TButton",
                        font=ui_font(9), padding=(14, 7))
        style.configure("SM.TButton",
                        font=ui_font(9), padding=(12, 6))
        style.configure("Add.TButton",
                        font=ui_font(10), padding=(14, 8), foreground=C["blue"])
        style.map("Add.TButton",
                  background=[("active", C["blue_light"])])
        style.configure("Del.TButton",
                        font=ui_font(9), padding=(8, 5), foreground=C["red_text"])

    # ────────── UI ──────────
    def _build_ui(self):
        outer = tk.Frame(self.root, bg=C["bg"])
        outer.pack(fill=tk.BOTH, expand=True)

        self.canvas = tk.Canvas(outer, bg=C["bg"], highlightthickness=0, bd=0)
        sbar = ttk.Scrollbar(outer, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=sbar.set)
        sbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.scroll_frame = tk.Frame(self.canvas, bg=C["bg"])
        self.canvas_window = self.canvas.create_window((0, 0), window=self.scroll_frame, anchor="nw")
        self.scroll_frame.bind("<Configure>", self._on_scroll_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.root.bind_all("<MouseWheel>", self._on_mousewheel, add="+")

        body = tk.Frame(self.scroll_frame, bg=C["bg"])
        body.pack(fill=tk.BOTH, expand=True, padx=20, pady=(16, 20))

        self._build_banner(body)
        self._build_status_cards(body)
        self._build_workarea(body)
        self._build_log(body)
        self._build_footer(body)

        self._log("欢迎使用 v4.0 多文件合并版 · 支持动态添加/删除多个 Excel 文件 🌸")

    # ─── Banner ───
    def _build_banner(self, parent):
        banner = tk.Frame(parent, bg=C["banner"], bd=0, highlightthickness=1,
                          highlightbackground=C["banner_deep"])
        banner.pack(fill=tk.X, pady=(0, 16))
        banner.config(padx=24, pady=18)

        tk.Label(banner, text=APP_TITLE,
                 font=ui_font(18, "bold"), bg=C["banner"], fg=C["text_inverse"]).pack(anchor="w")
        tk.Label(banner, text=f"{APP_SUBTITLE}  ·  {APP_VERSION}",
                 font=ui_font(9), bg=C["banner"], fg="#D3E4F7").pack(anchor="w", pady=(3, 0))
        tk.Label(banner,
                 text="离线处理｜多文件动态管理｜纵向追加 / 横向合并｜适合打包 EXE 分发",
                 font=ui_font(8), bg=C["banner"], fg="#E7F1FB").pack(anchor="w", pady=(6, 0))

        pill_row = tk.Frame(banner, bg=C["banner"])
        pill_row.pack(anchor="w", pady=(10, 0))
        for txt in ("离线可用", "多文件", "两种模式"):
            tk.Label(pill_row, text=f"  {txt}  ", font=ui_font(8, "bold"),
                     bg=C["banner_deep"], fg=C["text_inverse"], padx=4, pady=3).pack(side=tk.LEFT, padx=(0, 8))

    # ─── 状态卡片 ───
    def _build_status_cards(self, parent):
        cards = tk.Frame(parent, bg=C["bg"])
        cards.pack(fill=tk.X, pady=(0, 14))

        card_specs = [
            ("📂  文件列表",  self.file_count_var,       "可添加多个 Excel 文件",     C["blue"]),
            ("➡️  合并模式",  self.merge_mode,            "纵向追加 / 横向按列合并",   C["blue"]),
            ("💾  输出文件",  tk.StringVar(value="—"),     "生成后可直接打开",          C["purple"]),
        ]
        self.card_widgets = {}
        for i, (title, var, note, accent) in enumerate(card_specs):
            card = tk.Frame(cards, bg=C["card"], bd=1, relief=tk.FLAT,
                            highlightthickness=1, highlightbackground=C["border"])
            card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=4)
            card.config(padx=14, pady=10)

            tk.Label(card, text=title, font=ui_font(10, "bold"),
                     bg=C["card"], fg=C["text_primary"]).pack(anchor="w")
            val = tk.Label(card, textvariable=var, font=ui_font(12, "bold"),
                           bg=C["card"], fg=C["text_primary"])
            val.pack(anchor="w", pady=(6, 2))
            tk.Label(card, text=note, font=ui_font(8),
                     bg=C["card"], fg=C["text_muted"]).pack(anchor="w")

            self.card_widgets[title] = {"card": card, "val": val}

    # ─── 工作区 ───
    def _build_workarea(self, parent):
        work = tk.Frame(parent, bg=C["bg"])
        work.pack(fill=tk.BOTH, expand=True, pady=(0, 14))

        left = tk.Frame(work, bg=C["bg"])
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self._build_file_list(left)
        self._build_options(left)
        self._build_action_bar(left)

        right = tk.Frame(work, bg=C["bg"], width=300)
        right.pack(side=tk.RIGHT, fill=tk.Y)
        right.grid_propagate(False)
        self._build_guide(right)

    # ─── 文件列表 ───
    def _build_file_list(self, parent):
        hdr = tk.Frame(parent, bg=C["bg"])
        hdr.pack(fill=tk.X, pady=(8, 6))
        tk.Label(hdr, text="选择要合并的文件", font=ui_font(11, "bold"),
                 bg=C["bg"], fg=C["text_primary"]).pack(anchor="w")
        tk.Label(hdr, text="─" * 48, fg=C["border_light"], bg=C["bg"],
                 font=("Consolas", 7)).pack(anchor="w")

        list_bg = tk.Frame(parent, bg=C["card"], bd=1, relief=tk.FLAT,
                           highlightthickness=1, highlightbackground=C["border"])
        list_bg.pack(fill=tk.BOTH, expand=True, pady=(0, 8))
        list_bg.config(padx=8, pady=6)

        self.file_list_frame = tk.Frame(list_bg, bg=C["card"])
        self.file_list_frame.pack(fill=tk.BOTH, expand=True)

        self._show_empty_hint()

        btn_bar = tk.Frame(list_bg, bg=C["card"])
        btn_bar.pack(fill=tk.X, pady=(8, 2))
        ttk.Button(btn_bar, text="+  添加文件", style="Add.TButton",
                   command=self._add_file).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_bar, text="📂  浏览文件夹", style="SM.TButton",
                   command=self._add_folder).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_bar, text="清空所有", style="Del.TButton",
                   command=self._clear_all).pack(side=tk.LEFT)

    def _show_empty_hint(self):
        for w in self.file_list_frame.winfo_children():
            w.destroy()
        hint = tk.Frame(self.file_list_frame, bg=C["card_alt"], bd=0)
        hint.pack(fill=tk.X, ipady=20)
        hint.config(padx=14)
        tk.Label(hint, text="暂无文件，点击上方「+ 添加文件」或「浏览文件夹」开始",
                 font=ui_font(9), bg=C["card_alt"], fg=C["text_muted"]).pack(anchor="w")
        tk.Label(hint, text="支持 .xlsx 和 .xls 格式，可一次添加多个文件",
                 font=ui_font(8), bg=C["card_alt"], fg=C["text_muted"]).pack(anchor="w", pady=(4, 0))

    def _render_file_list(self):
        for w in self.file_list_frame.winfo_children():
            w.destroy()
        if not self.file_items:
            self._show_empty_hint()
            self._update_status()
            return

        for idx, item in enumerate(self.file_items):
            row = tk.Frame(self.file_list_frame, bg=C["bg"], bd=1, relief=tk.FLAT,
                           highlightthickness=1, highlightbackground=C["border"])
            row.pack(fill=tk.X, pady=(0, 4))
            row.config(pady=8, padx=8)

            tk.Label(row, text=f"[{idx + 1}]", font=ui_font(9, "bold"),
                     bg=C["bg"], fg=C["blue"], width=4, anchor="w").pack(side=tk.LEFT, padx=(0, 4))

            info = tk.Frame(row, bg=C["bg"])
            info.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            tk.Label(info, text=item.name, font=ui_font(9, "bold"),
                     bg=C["bg"], fg=C["text_primary"]).pack(anchor="w")
            tk.Label(info, text=item.path, font=ui_font(7),
                     bg=C["bg"], fg=C["text_muted"]).pack(anchor="w")

            tk.Label(row, text=item.status, font=ui_font(8),
                     bg=C["bg"], fg=C["green_text"]).pack(side=tk.LEFT, padx=(0, 8))

            if idx > 0:
                ttk.Button(row, text="↑", style="SM.TButton",
                           command=lambda i=idx: self._move_up(i)).pack(side=tk.LEFT, padx=(0, 2))
            if idx < len(self.file_items) - 1:
                ttk.Button(row, text="↓", style="SM.TButton",
                           command=lambda i=idx: self._move_down(i)).pack(side=tk.LEFT, padx=(0, 2))
            ttk.Button(row, text="✕", style="Del.TButton",
                       command=lambda i=idx: self._remove_file(i)).pack(side=tk.LEFT)

        self._update_status()

    # ─── 合并选项 ───
    def _build_options(self, parent):
        hdr = tk.Frame(parent, bg=C["bg"])
        hdr.pack(fill=tk.X, pady=(8, 6))
        tk.Label(hdr, text="合并选项", font=ui_font(11, "bold"),
                 bg=C["bg"], fg=C["text_primary"]).pack(anchor="w")
        tk.Label(hdr, text="─" * 48, fg=C["border_light"], bg=C["bg"],
                 font=("Consolas", 7)).pack(anchor="w")

        opt_box = tk.Frame(parent, bg=C["card_alt"], bd=1, relief=tk.FLAT,
                           highlightthickness=1, highlightbackground=C["border"])
        opt_box.pack(fill=tk.X, pady=(0, 8))
        opt_box.config(padx=14, pady=12)

        mode_row = tk.Frame(opt_box, bg=C["card_alt"])
        mode_row.pack(fill=tk.X, pady=(0, 8))
        tk.Label(mode_row, text="合并模式：", font=ui_font(9, "bold"),
                 bg=C["card_alt"], fg=C["text_primary"], width=12, anchor="w").pack(side=tk.LEFT)

        ttk.Radiobutton(mode_row, text="纵向追加（stack）", variable=self.merge_mode,
                        value="stack").pack(side=tk.LEFT, padx=(0, 16))
        ttk.Radiobutton(mode_row, text="横向合并（join）", variable=self.merge_mode,
                        value="join").pack(side=tk.LEFT)

        key_row = tk.Frame(opt_box, bg=C["card_alt"])
        key_row.pack(fill=tk.X, pady=(4, 0))
        tk.Label(key_row, text="合并键(列名)：", font=ui_font(9, "bold"),
                 bg=C["card_alt"], fg=C["text_primary"], width=12, anchor="w").pack(side=tk.LEFT)
        tk.Entry(key_row, textvariable=self.join_key, font=ui_font(9),
                 bg="#FFFFFF", fg=C["text_primary"], width=20, relief=tk.SOLID, bd=1).pack(side=tk.LEFT, padx=(0, 8))
        tk.Label(key_row, text="（仅横向合并时用，模糊匹配）", font=ui_font(8),
                 bg=C["card_alt"], fg=C["text_muted"]).pack(side=tk.LEFT)

        # 指定列过滤（仅横向合并生效）
        col_row = tk.Frame(opt_box, bg=C["card_alt"])
        col_row.pack(fill=tk.X, pady=(4, 0))
        tk.Label(col_row, text="指定列（可选）：", font=ui_font(9, "bold"),
                 bg=C["card_alt"], fg=C["text_primary"], width=12, anchor="w").pack(side=tk.LEFT)
        tk.Entry(col_row, textvariable=self.column_filter, font=ui_font(9),
                 bg="#FFFFFF", fg=C["text_primary"], width=30, relief=tk.SOLID, bd=1).pack(side=tk.LEFT, padx=(0, 8))
        tk.Label(col_row, text="（仅横向合并，逗号分隔，留空=全部）", font=ui_font(8),
                 bg=C["card_alt"], fg=C["text_muted"]).pack(side=tk.LEFT)

        mode_hint = tk.Frame(opt_box, bg=C["card_alt"])
        mode_hint.pack(fill=tk.X, pady=(8, 0))
        tk.Label(mode_hint,
                 text="纵向：多个文件行追加合并（同结构表）  |  横向：按列名关联合并（不同维度表）",
                 font=ui_font(8), bg=C["card_alt"], fg=C["text_muted"]).pack(anchor="w")

    # ─── 操作栏 ───
    def _build_action_bar(self, parent):
        bar = tk.Frame(parent, bg=C["action_bg"], bd=1, relief=tk.FLAT, highlightthickness=1,
                       highlightbackground=C["blue_border"])
        bar.pack(fill=tk.X, pady=(4, 0))
        bar.config(padx=18, pady=14)

        txt_col = tk.Frame(bar, bg=C["action_bg"])
        txt_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tk.Label(txt_col, text="⚡ 执行操作", font=ui_font(11, "bold"),
                 bg=C["action_bg"], fg=C["blue"]).pack(anchor="w")
        tk.Label(txt_col, text="确认文件和选项无误后，点击主按钮开始合并",
                 font=ui_font(8), bg=C["action_bg"], fg=C["text_secondary"]).pack(anchor="w", pady=(3, 0))

        btn_row = tk.Frame(bar, bg=C["action_bg"])
        btn_row.pack(side=tk.RIGHT)
        self.merge_btn = ttk.Button(btn_row, text="▶  一键合并", style="Primary.TButton",
                                    command=self._start_merge, width=14)
        self.merge_btn.pack(side=tk.LEFT)

    # ─── 说明面板 ───
    def _build_guide(self, parent):
        card = tk.Frame(parent, bg=C["card"], bd=1, relief=tk.FLAT, highlightthickness=1,
                        highlightbackground=C["border"])
        card.pack(fill=tk.BOTH, expand=True)
        card.config(padx=16, pady=14)

        tk.Label(card, text="📖 使用说明", font=ui_font(11, "bold"),
                 bg=C["card"], fg=C["text_primary"]).pack(anchor="w")
        tk.Label(card, text="─" * 32, fg=C["border_light"], bg=C["card"],
                 font=("Consolas", 7)).pack(anchor="w", pady=(4, 10))

        self._step(card, "1", "添加文件", "点击「+ 添加文件」选择多个 Excel，或「浏览文件夹」批量导入")
        self._step(card, "2", "选择模式", "纵向追加：多个同结构表的行追加\n横向合并：按列名关联不同表的数据")
        self._step(card, "3", "确认合并", "确认选项后，点击主按钮执行合并。合并键在横向模式下使用。")
        self._step(card, "4", "查看结果", "合并完成后可直接打开文件夹 / 结果文件")

        tk.Label(card, text="─" * 32, fg=C["border_light"], bg=C["card"],
                 font=("Consolas", 7)).pack(anchor="w", pady=(10, 10))
        tk.Label(card, text="✨ 工具特性", font=ui_font(10, "bold"),
                 bg=C["card"], fg=C["blue"]).pack(anchor="w")

        for f in [
            "✓  动态添加/删除文件，可调整顺序",
            "✓  支持纵向追加和横向合并两种模式",
            "✓  自动处理表头行定位",
            "✓  全程离线，无需联网",
            "✓  适合打包 EXE 后直接分发",
        ]:
            tk.Label(card, text=f, font=ui_font(9), bg=C["card"], fg=C["text_secondary"],
                     justify="left").pack(anchor="w", pady=(3, 0))

    def _step(self, parent, num, title, desc):
        row = tk.Frame(parent, bg=C["card"])
        row.pack(fill=tk.X, pady=(0, 8))
        badge = tk.Label(row, text=num, font=ui_font(9, "bold"), bg=C["blue"], fg="#FFFFFF",
                         width=2, relief=tk.FLAT)
        badge.pack(side=tk.LEFT, padx=(0, 10), ipady=3)
        col = tk.Frame(row, bg=C["card"])
        col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tk.Label(col, text=title, font=ui_font(9, "bold"), bg=C["card"], fg=C["text_primary"]).pack(anchor="w")
        tk.Label(col, text=desc, font=ui_font(8), bg=C["card"], fg=C["text_secondary"],
                 wraplength=220, justify="left").pack(anchor="w", pady=(1, 0))

    # ─── 日志 ───
    def _build_log(self, parent):
        frame = tk.Frame(parent, bg=C["card"], bd=1, relief=tk.FLAT, highlightthickness=1,
                         highlightbackground=C["border"])
        frame.pack(fill=tk.BOTH, expand=True, pady=(0, 0))
        frame.config(padx=14, pady=10)

        hdr = tk.Frame(frame, bg=C["card"])
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text="⚙ 运行日志", font=ui_font(10, "bold"),
                 bg=C["card"], fg=C["text_primary"]).pack(side=tk.LEFT)
        ttk.Button(hdr, text="清空", style="SM.TButton", command=self._clear_log).pack(side=tk.RIGHT)

        self.log_text = scrolledtext.ScrolledText(
            frame, height=12, font=mono_font(8),
            bg=C["term_bg"], fg="#8FA4BF",
            insertbackground="#FFFFFF",
            wrap=tk.WORD, relief=tk.FLAT,
            padx=12, pady=8, bd=0
        )
        self.log_text.pack(fill=tk.BOTH, expand=True, pady=(8, 0))

        bottom = tk.Frame(frame, bg=C["card"])
        bottom.pack(fill=tk.X, pady=(10, 0))

        st_bar = tk.Frame(bottom, bg=C["blue_light"], padx=12, pady=8, bd=1, relief=tk.FLAT,
                          highlightthickness=1, highlightbackground=C["blue_border"])
        st_bar.pack(fill=tk.X, pady=(0, 6))
        tk.Label(st_bar, text="状态：", font=ui_font(9, "bold"),
                 bg=C["blue_light"], fg=C["blue"]).pack(side=tk.LEFT)
        self.status_value_label = tk.Label(st_bar, textvariable=self.status_text,
                                           font=ui_font(9, "bold"), bg=C["blue_light"], fg=C["blue"])
        self.status_value_label.pack(side=tk.LEFT, padx=(4, 0))

        sm_bar = tk.Frame(bottom, bg=C["card_alt"], padx=12, pady=8, bd=1, relief=tk.FLAT,
                          highlightthickness=1, highlightbackground=C["border"])
        sm_bar.pack(fill=tk.X)
        tk.Label(sm_bar, text="摘要：", font=ui_font(9, "bold"),
                 bg=C["card_alt"], fg=C["text_primary"]).pack(side=tk.LEFT)
        tk.Label(sm_bar, textvariable=self.summary_text, font=ui_font(9), bg=C["card_alt"],
                 fg=C["text_secondary"], wraplength=900, justify="left").pack(
            side=tk.LEFT, padx=(4, 0), fill=tk.X, expand=True)

        rb = tk.Frame(bottom, bg=C["card"])
        rb.pack(fill=tk.X, pady=(8, 0))
        self.open_out_btn = ttk.Button(rb, text="📂 打开输出文件夹", style="SM.TButton",
                                       command=self._open_output_folder, width=16)
        self.open_out_btn.pack(side=tk.LEFT, padx=(0, 6))
        self.open_file_btn = ttk.Button(rb, text="📄 打开结果文件", style="SM.TButton",
                                        command=self._open_output_file, width=14)
        self.open_file_btn.pack(side=tk.LEFT)

    # ─── 底栏 ───
    def _build_footer(self, parent):
        tk.Label(parent,
                 text=f"{APP_TITLE}  {APP_VERSION}  ·  离线运行 · 本地处理 · 可打包 EXE 分发",
                 font=ui_font(8), bg=C["bg"], fg=C["text_muted"], justify="center").pack(anchor="center", pady=(14, 0))

    # ────────── 滚动 ──────────
    def _on_scroll_configure(self, event=None):
        if hasattr(self, "canvas"):
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        if hasattr(self, "canvas") and hasattr(self, "canvas_window"):
            self.canvas.itemconfigure(self.canvas_window, width=event.width)

    def _on_mousewheel(self, event):
        """滚动处理：日志区域滚到底/顶后继续滚动外层页面"""
        try:
            steps = int(-event.delta / 40) or (-3 if event.delta > 0 else 3)
            
            # 检查是否在日志区域
            if hasattr(self, "log_text"):
                px = self.root.winfo_pointerx()
                py = self.root.winfo_pointery()
                widget = self.root.winfo_containing(px, py)
                if widget:
                    w = widget
                    while w is not None:
                        if w is self.log_text:
                            # 检查日志是否还能继续滚动
                            top, bottom = self.log_text.yview()
                            # 向下滚且已到底，或向上滚且已到顶，则让外层处理
                            if (steps > 0 and bottom >= 0.999) or (steps < 0 and top <= 0.001):
                                break  # 让外层 canvas 处理
                            self.log_text.yview_scroll(steps, "units")
                            return
                        w = w.master
            
            # 滚动外层 canvas
            if hasattr(self, "canvas"):
                self.canvas.yview_scroll(steps, "units")
        except Exception:
            pass

    # ────────── 文件操作 ──────────
    def _add_file(self):
        paths = filedialog.askopenfilenames(
            filetypes=[("Excel files", "*.xlsx *.xls")],
            title="选择要合并的 Excel 文件")
        if not paths:
            return
        for p in paths:
            self.file_items.append(FileItem(path=p, name=os.path.basename(p)))
        self._render_file_list()
        self._log(f"已添加 {len(paths)} 个文件")

    def _add_folder(self):
        folder = filedialog.askdirectory(title="选择包含 Excel 文件的文件夹")
        if not folder:
            return
        folder = Path(folder)
        files = sorted(folder.glob("*.xls*"))
        if not files:
            messagebox.showwarning("提示", f"文件夹中没有找到 Excel 文件：\n{folder}")
            return
        for f in files:
            self.file_items.append(FileItem(path=str(f), name=f.name))
        self._render_file_list()
        self._log(f"已从文件夹添加 {len(files)} 个文件: {folder}")

    def _clear_all(self):
        if not self.file_items:
            return
        if messagebox.askyesno("确认", f"确认清空所有 {len(self.file_items)} 个文件？"):
            self.file_items.clear()
            self._render_file_list()
            self._log("已清空所有文件")

    def _move_up(self, idx):
        if idx > 0:
            self.file_items[idx], self.file_items[idx - 1] = self.file_items[idx - 1], self.file_items[idx]
            self._render_file_list()

    def _move_down(self, idx):
        if idx < len(self.file_items) - 1:
            self.file_items[idx], self.file_items[idx + 1] = self.file_items[idx + 1], self.file_items[idx]
            self._render_file_list()

    def _remove_file(self, idx):
        self.file_items.pop(idx)
        self._render_file_list()

    # ────────── 状态更新 ──────────
    def _update_status(self):
        if not self.file_items:
            self.file_count_var.set("0 个文件")
        else:
            self.file_count_var.set(f"✓ {len(self.file_items)} 个文件")
        ready = len(self.file_items) >= 2
        if hasattr(self, "merge_btn"):
            self.merge_btn.config(state=tk.NORMAL if ready else tk.DISABLED)

    # ────────── 自动扫描 ──────────
    def _auto_scan_defaults(self):
        folder = DEFAULT_FOLDER
        if not folder.exists():
            return
        files = sorted(folder.glob("*.xls*"))
        if len(files) >= 2:
            for f in files:
                self.file_items.append(FileItem(path=str(f), name=f.name))
            self._render_file_list()
            self._log(f"自动扫描到 {len(files)} 个文件: {folder}")
            self._set_status("已自动识别默认目录", C["blue"])

    # ────────── 辅助 ──────────
    def _set_status(self, msg, color=C["text_secondary"]):
        self.status_text.set(msg)
        if hasattr(self, "status_value_label"):
            self.status_value_label.config(fg=color)

    def _set_summary(self, msg):
        self.summary_text.set(msg)

    # ────────── 合并 ──────────
    def _start_merge(self):
        if len(self.file_items) < 2:
            messagebox.showwarning("提示", "至少需要 2 个文件才能合并。")
            return

        out_path = self.output_path.get().strip() or str(DEFAULT_FOLDER / OUTPUT_NAME)
        self._log("开始合并...")
        self._set_status("正在处理...", C["orange_text"])
        self._set_summary(f"正在合并 {len(self.file_items)} 个文件")
        if self.merge_btn:
            self.merge_btn.config(state=tk.DISABLED)

        threading.Thread(target=self._do_merge, args=(out_path,), daemon=True).start()

    def _find_col(self, df, candidates):
        cols = [str(c) for c in df.columns]
        for c in candidates:
            for col in cols:
                if c in col:
                    return df.columns[cols.index(col)]
        return None

    def _pick_columns(self, df, key_col, filter_text):
        """根据逗号分隔的关键词，从 DataFrame 中选择列（模糊匹配）。"""
        filter_cols = [str(c).strip() for c in filter_text.split(",") if str(c).strip()]
        keep = [key_col]
        for fc in filter_cols:
            if not fc:
                continue
            for col in df.columns:
                col_name = str(col) if col is not None else ""
                if col_name and fc in col_name and col != key_col:
                    keep.append(col)
                    break
        self._log(f"  → 列过滤选定：{keep[1:]}", "INFO")
        return keep

    def _write_formatted_excel(self, result_df, out_path):
        """输出精美样式的 Excel，对齐参考文件 3.最终需要的表格.xlsx"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "汇总成绩"

        num_cols = len(result_df.columns)
        end_col = openpyxl.utils.get_column_letter(num_cols)

        # 参考文件标准列宽
        col_widths = {
            "A": 7.2, "B": 17.1, "C": 22.6, "D": 14.4, "E": 14.5,
            "F": 16.9, "G": 12.8, "H": 10.1, "I": 9.4, "J": 9.0,
        }

        # 样式定义
        title_fill = PatternFill(start_color="78A9F2", end_color="78A9F2", fill_type="solid")
        title_font = Font(name="Microsoft YaHei", size=14, bold=True, color="FFFFFF")
        title_align = Alignment(horizontal="center", vertical="center")

        header_fill = PatternFill(start_color="CBDFFE", end_color="CBDFFE", fill_type="solid")
        header_font = Font(name="Microsoft YaHei", size=10, bold=True, color="1F2D3D")
        header_align = Alignment(horizontal="center", vertical="center")

        data_font = Font(name="Microsoft YaHei", size=10)
        data_align = Alignment(horizontal="center", vertical="center")
        odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        even_fill = PatternFill(start_color="E7F1FA", end_color="E7F1FA", fill_type="solid")

        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Row1: 标题行（合并、浅蓝底、白字、居中）
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        ws.row_dimensions[1].height = 26
        cell_a1 = ws.cell(row=1, column=1, value="平时成绩情况一览表")
        cell_a1.font = title_font
        cell_a1.fill = title_fill
        cell_a1.alignment = title_align

        # Row2: 表头行（浅蓝底、黑粗体、居中、边框）
        headers = list(result_df.columns)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
        ws.row_dimensions[2].height = 22

        # Row3+: 数据行（交替底色、居中、边框）
        for row_idx, (_, row_data) in enumerate(result_df.iterrows(), start=3):
            fill = odd_fill if row_idx % 2 == 1 else even_fill
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.fill = fill
                cell.alignment = data_align
                cell.border = border
            # B列(学号)设为文本格式
            ws.cell(row=row_idx, column=2).number_format = "@"
            ws.row_dimensions[row_idx].height = 18

        # 列宽
        for i in range(1, num_cols + 1):
            letter = openpyxl.utils.get_column_letter(i)
            ws.column_dimensions[letter].width = col_widths.get(letter, 14)

        # 冻结窗格 / 筛选 / 隐藏网格
        ws.freeze_panes = "C3"
        ws.sheet_view.showGridLines = False
        ws.auto_filter.ref = f"A2:{end_col}{len(result_df) + 2}"

        out_path = Path(str(out_path)) if not isinstance(out_path, Path) else out_path
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(out_path))

    def _read_excel_smart(self, file_path, key_hint=None):
        """智能读取 Excel：自动扫描表头行"""
        df = pd.read_excel(file_path, header=0)
        # 如果不需要找 key_hint，直接返回
        if key_hint is None:
            df = df.dropna(how="all").reset_index(drop=True)
            return df
        # 尝试找 key_hint 是否在 header=0 中
        if self._find_col(df, [key_hint]):
            df = df.dropna(subset=[df.columns[0]], how="all").reset_index(drop=True)
            return df
        # 扫描找表头
        df_raw = pd.read_excel(file_path, header=None)
        self._log(f"  → 扫描表头（共 {df_raw.shape[0]} 行）", "INFO")
        for r in range(min(df_raw.shape[0], 20)):
            row_vals = df_raw.iloc[r].astype(str).tolist()
            if any(key_hint in str(v) for v in row_vals):
                self._log(f"  → 找到表头在第 {r+1} 行", "OK")
                df = pd.read_excel(file_path, header=r)
                df = df.dropna(subset=[df.columns[0]], how="all").reset_index(drop=True)
                return df
        # 兜底
        df = pd.read_excel(file_path, header=0)
        df = df.dropna(how="all").reset_index(drop=True)
        return df

    def _do_merge(self, out_path):
        try:
            mode = self.merge_mode.get()
            self._log(f"合并模式：{'纵向追加' if mode == 'stack' else '横向合并'}")
            self._log(f"文件数量：{len(self.file_items)}")

            if mode == "stack":
                # 纵向追加：不应用列过滤
                self._log("纵向模式")
                dfs = []
                for i, item in enumerate(self.file_items):
                    self._log(f"[{i+1}] 读取: {item.name}")
                    # header=1 跳过第 0 行（"综合成绩"标题行），从第 1 行的真正表头读取
                    df = pd.read_excel(item.path, header=1)
                    df = df.dropna(how="all").reset_index(drop=True)
                    self._log(f"  → {len(df)} 行，{len(df.columns)} 列（已跳过标题行）")
                    dfs.append(df)
                result = pd.concat(dfs, ignore_index=True)
                result.reset_index(drop=True, inplace=True)
                self._log(f"纵向合并完成：{len(result)} 行，{len(result.columns)} 列", "OK")
                # 加序号列（如果已存在先删再插，确保在第一列）
                if "序号" in result.columns:
                    result = result.drop(columns=["序号"])
                result.insert(0, "序号", range(1, len(result) + 1))
                self._log(f"纵向合并完成：{len(result)} 行，{len(result.columns)} 列", "OK")

            else:
                # ✅ 横向合并：应用列过滤
                join_key_hint = self.join_key.get().strip() or "学号/工号"
                filter_text = (self.column_filter.get() or "").strip()
                self._log(f"合并键关键词：{join_key_hint}")
                if filter_text:
                    self._log(f"列过滤关键词：{filter_text}")

                result = None
                for i, item in enumerate(self.file_items):
                    self._log(f"[{i+1}] 读取: {item.name}")
                    df = self._read_excel_smart(item.path, key_hint=join_key_hint)
                    key_col = self._find_col(df, [join_key_hint])
                    if key_col is None:
                        cols_preview = ", ".join([str(c) for c in df.columns[:8]])
                        self._log(f"  ⚠ 找不到包含 '{join_key_hint}' 的列：{cols_preview}...", "WARN")
                        continue

                    df[key_col] = df[key_col].astype(str).str.replace(".0", "", regex=False).str.strip()

                    if result is None:
                        # 第一个文件：如果指定了列过滤，只取指定列
                        if filter_text:
                            keep = self._pick_columns(df, key_col, filter_text)
                            if len(keep) <= 1:
                                self._log(f"  ⚠ 没有匹配到列，使用全部列作为 fallback", "WARN")
                                result = df.copy()
                            else:
                                result = df[keep].copy()
                                self._log(f"  → 初始表（已过滤）：{len(result)} 行 × {len(result.columns)} 列", "OK")
                        else:
                            result = df.copy()
                            self._log(f"  → 初始表：{len(result)} 行 × {len(result.columns)} 列", "OK")
                    else:
                        # 后续文件：统一 key 列类型
                        key_in_result = self._find_col(result, [join_key_hint])
                        if key_in_result is None:
                            self._log(f"  ⚠ 结果中没有合并键列，跳过", "WARN")
                            continue
                        result[key_in_result] = result[key_in_result].astype(str).str.replace(".0", "", regex=False).str.strip()

                        # 列过滤
                        if filter_text:
                            keep = self._pick_columns(df, key_col, filter_text)
                            extra = [c for c in keep if c not in result.columns and c != key_col]
                        else:
                            extra = [c for c in df.columns if c not in result.columns and c != key_col]

                        if not extra:
                            self._log(f"  → 此文件无新列可合并，跳过", "WARN")
                            continue

                        merge_df = df[[key_col] + extra].copy()
                        result = result.merge(merge_df, left_on=key_in_result, right_on=key_col, how="outer")
                        # 删除重复的 key 列
                        if key_col != key_in_result and key_col in result.columns:
                            result = result.drop(columns=[key_col])
                        self._log(f"  → 新增 {len(extra)} 列，合并后 {len(result)} 行 × {len(result.columns)} 列", "OK")

                if result is None or len(result) == 0:
                    raise ValueError("找不到合并键列，横向合并失败。")
                # 清理 unnamed 列
                result = result.loc[:, ~result.columns.str.contains('^Unnamed', na=False)]
                result.reset_index(drop=True, inplace=True)
                # 加序号列（先删旧再插入到第一列）
                if "序号" in result.columns:
                    result = result.drop(columns=["序号"])
                result.insert(0, "序号", range(1, len(result) + 1))
                self._log(f"横向合并完成：{len(result)} 行 × {len(result.columns)} 列（已加序号）", "OK")

            # 输出：按模式和时间戳生成不重复文件
            import datetime
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            suffix = "横向合并" if mode == "join" else "纵向追加"
            out_dir = Path(out_path).parent
            out_dir.mkdir(parents=True, exist_ok=True)
            out_file = out_dir / f"合并结果_{suffix}_{ts}.xlsx"
            self._log(f"输出模式：{suffix}，文件名自动加时间戳")

            # 两种模式都使用精美排版
            self._write_formatted_excel(result, str(out_file))
            self._log(f"已保存：{out_file}", "OK")
            self.output_path.set(str(out_file))
            self._set_status("合并成功！", C["green_text"])
            self._set_summary(f"完成：{len(result)} 行 × {len(result.columns)} 列")
            messagebox.showinfo("合并完成",
                f"成功合并 {len(self.file_items)} 个文件！\n\n"
                f"结果：{len(result)} 行 × {len(result.columns)} 列\n\n"
                f"输出：\n{out_file}")

        except Exception as e:
            self._log(f"合并失败：{e}", "ERROR")
            self._log(traceback.format_exc(), "ERROR")
            self._set_status("合并失败", C["red_text"])
            self._set_summary("处理失败，请查看日志")
            messagebox.showerror("错误", f"合并失败：\n\n{e}")
        finally:
            if self.merge_btn:
                self.merge_btn.config(state=tk.NORMAL)

    # ────────── 打开文件/文件夹 ──────────
    def _open_output_folder(self):
        out = self.output_path.get().strip()
        if not out:
            messagebox.showwarning("提示", "请先生成合并结果。")
            return
        folder = Path(out).parent
        if folder.exists():
            try:
                os.startfile(folder)
            except Exception as e:
                messagebox.showerror("错误", f"打开文件夹失败:\n{e}")

    def _open_output_file(self):
        out = self.output_path.get().strip()
        if not out:
            messagebox.showwarning("提示", "请先生成合并结果。")
            return
        f = Path(out)
        if f.exists():
            try:
                os.startfile(f)
            except Exception as e:
                messagebox.showerror("错误", f"打开文件失败:\n{e}")

    # ────────── 日志操作 ──────────
    def _log(self, msg, level="INFO"):
        prefix = {"INFO": "[INFO]", "OK": "[OK] ✓", "WARN": "[WARN]", "ERROR": "[ERROR]"}.get(level, "[INFO]")
        if self.log_text:
            self.log_text.insert(tk.END, f"{prefix} {msg}\n")
            self.log_text.see(tk.END)
            self.root.update_idletasks()

    def _clear_log(self):
        if self.log_text:
            self.log_text.delete(1.0, tk.END)
            self._log("日志已清空。")


def main():
    enable_high_dpi()
    root = tk.Tk()
    scale = detect_ui_scale()
    try:
        root.tk.call("tk", "scaling", scale)
    except Exception:
        pass
    MultiMergeGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
