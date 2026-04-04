#!/usr/bin/env python3
"""
学习通成绩汇总工具（发布级 GUI v3.1）
- 脱机运行，无需联网
- 将班级成绩表和期末考试成绩表合并为汇总格式
- 自动识别常见文件名，也支持手动选择
- 全新视觉排版（功能不变，界面升级）
- 适合作为 Windows EXE 直接分发
- 打包为 Windows EXE:
  py -m PyInstaller --onefile --windowed --name 成绩汇总工具 score_merger_gui.py
"""

import os
import sys
import ctypes
import traceback
import threading
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

APP_TITLE = "学习通成绩汇总工具"
APP_VERSION = "v3.1 发布版"
APP_SUBTITLE = "离线成绩汇总桌面工具"
DEFAULT_FOLDER = Path.home() / "Desktop" / "表格输入输出" / "表格导入导出"
CLASS_PATTERNS = ["1.学习通-班级成绩.xlsx", "学习通-班级成绩.xlsx", "班级成绩.xlsx"]
EXAM_PATTERNS = ["2.学习通-期末考试客观题试卷-考.xlsx", "学习通-期末考试客观题试卷-考.xlsx", "期末考试成绩.xlsx"]
OUTPUT_NAME = "汇总成绩.xlsx"
UI_SCALE = 1.0

# ============================================================
#  配色系统（Design Token）
# ============================================================
C = {
    # 背景
    "bg":            "#F5F7FA",
    "card":          "#FFFFFF",
    "card_alt":      "#FAFBFD",
    "banner":        "#1B6AB0",
    "banner_deep":   "#154E83",
    "banner_light":  "#E8F1FA",
    "action_bg":     "#F0F6FF",
    "term_bg":       "#141D2A",

    # 文字
    "text_primary":  "#1A2636",
    "text_secondary":"#5A6B7E",
    "text_muted":    "#8B97A7",
    "text_link":     "#2078BC",
    "text_inverse":  "#FFFFFF",

    # 主题色
    "blue":          "#2B6CB0",
    "blue_hover":    "#245E99",
    "blue_light":    "#EBF4FF",
    "blue_border":   "#BFD4ED",
    "purple":        "#7C5CC4",

    # 状态色
    "green_bg":      "#E9F7EF",
    "green_border":  "#81E2A9",
    "green_text":    "#187A4B",
    "orange_bg":     "#FFF6EC",
    "orange_border": "#F5C579",
    "orange_text":   "#C77600",
    "red_bg":        "#FEECED",
    "red_border":    "#F5A5A8",
    "red_text":      "#C43A43",

    # 分隔/边框
    "border":        "#E4E8ED",
    "border_light":  "#EEF1F5",
    "shadow":        "#DCE5EF",
}

CLASS_KEYWORDS = [
    ("班级成绩",),
    ("学习通", "班级", "成绩"),
]

EXAM_KEYWORDS = [
    ("期末考试",),
    ("期末", "考试"),
    ("客观题", "试卷"),
    ("学习通", "期末", "考试"),
]

# ============================================================
#  字体工具（支持高DPI自适应）
# ============================================================
def ui_font(size, weight="normal", family="Microsoft YaHei UI"):
    scaled = max(8, int(round(size * UI_SCALE)))
    return (family, scaled, weight)

def ui_font_fallback(size, weight="normal", family="Microsoft YaHei"):
    scaled = max(8, int(round(size * UI_SCALE)))
    return (family, scaled, weight)

def mono_font(size=9):
    scaled = max(9, int(round(size * UI_SCALE)))
    return ("Cascadia Mono", scaled)

def enable_high_dpi():
    if not sys.platform.startswith("win"):
        return
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass

def detect_ui_scale():
    if not sys.platform.startswith("win"):
        return 1.04
    try:
        user32 = ctypes.windll.user32
        gdi32 = ctypes.windll.gdi32
        hwnd = user32.GetDesktopWindow()
        dc = user32.GetDC(hwnd)
        try:
            dpi_x = gdi32.GetDeviceCaps(dc, 88)
        finally:
            user32.ReleaseDC(hwnd, dc)
        scale = dpi_x / 96.0
        scale = max(1.0, min(scale * 1.04, 1.5))
        return round(scale, 2)
    except Exception:
        return 1.2

def get_icon_path() -> str | None:
    """获取图标路径，优先使用 PyInstaller 打包后资源，回退到脚本目录。"""
    if getattr(sys, "frozen", False):
        base = Path(sys._MEIPASS)
    else:
        base = Path(__file__).parent
    for name in ["score_merger_icon.ico", "icon.ico"]:
        p = base / name
        if p.exists():
            return str(p)
    return None

def set_window_icon(root: tk.Tk):
    """设置窗口图标（EXE 资源 + wm_iconphoto 双保险）。"""
    icon = get_icon_path()
    if not icon:
        return
    if not sys.platform.startswith("win"):
        return
    try:
        root.iconbitmap(icon)
    except Exception:
        pass
    try:
        img = tk.PhotoImage(file=icon)
        root.iconphoto(True, img)
        _icon_ref = img  # prevent GC
    except Exception:
        pass

def configure_window(root: tk.Tk):
    try:
        root.update_idletasks()
        sw = root.winfo_screenwidth()
        sh = root.winfo_screenheight()
        tw = min(1180, max(980, sw - 120))
        th = min(820, max(700, sh - 120))
        mw = min(960, tw)
        mh = min(680, th)
        px = max((sw - tw) // 2, 20)
        py = max((sh - th) // 2, 20)
        root.geometry(f"{tw}x{th}+{px}+{py}")
        root.minsize(mw, mh)
    except Exception:
        root.geometry("1120x780")
        root.minsize(960, 680)

# ============================================================
#  圆角卡片（Canvas绘制）
# ============================================================
CARD_W = 220
CARD_H = 82
RADIUS = 10

def _create_rounded_rect(canvas, x0, y0, x1, y1, r, bg, outline=None, width=0):
    canvas.create_rectangle(x0+r, y0, x1-r, y1, fill=bg, outline="", width=0)
    canvas.create_rectangle(x0, y0+r, x1, y1-r, fill=bg, outline="", width=0)
    for dx0, dy0 in [(x0, y0), (x1, y0), (x0, y1), (x1, y1)]:
        canvas.create_oval(dx0, dy0, dx0 + 2*r*(1 if dx0==x0 else -1), dy0 + 2*r*(1 if dy0==y0 else -1),
                           fill=bg, outline="")
    if outline and width:
        canvas.create_line(x0+r, y0, x1-r, y0, fill=outline, width=width)
        canvas.create_line(x0+r, y1, x1-r, y1, fill=outline, width=width)
        canvas.create_line(x0, y0+r, x0, y1-r, fill=outline, width=width)
        canvas.create_line(x1, y0+r, x1, y1-r, fill=outline, width=width)
        for dx0, dy0, sx, sy in [(x0+r, y0, -1, 1), (x1-r, y0, 1, 1), (x0+r, y1, -1, -1), (x1-r, y1, 1, -1)]:
            canvas.create_arc(dx0, dy0, dx0 + 2*rx*(1), dy0 + 2*ry*(1), style=tk.ARC, outline=outline, width=width)
        # arc 简化替代：不画弧，只画四边线（视觉已足够）

def _card_bg_bg(canvas, x0, y0, w, h, r, bg, accent, border):
    """绘制带左侧色条的圆角卡片背景"""
    # 背景
    canvas.create_rectangle(x0, y0, x0+w, y0+h, fill=bg, outline=border, width=1)
    # 左侧色条
    canvas.create_rectangle(x0, y0+6, x0+5, y0+h-6, fill=accent, outline="")
    # 圆角覆盖
    for dx, dy, s in [(x0, y0, 1), (x0+w, y0, 1), (x0, y0+h, -1), (x0+w, y0+h, -1)]:
        canvas.create_oval(dx - r*s, dy - r*s, dx + r*(-s) + 2*r if s<0 else dx - r*s,
                          dy - r*s, dy + r*(-s) + 2*r if s<0 else dy - r*s, fill=bg)

# ============================================================
#  主应用
# ============================================================
class ScoreMergerGUI:
    def __init__(self, root):
        global UI_SCALE
        UI_SCALE = detect_ui_scale()

        self.root = root
        self.root.title(f" {APP_TITLE} {APP_VERSION}")
        set_window_icon(self.root)
        configure_window(self.root)
        self.root.configure(bg=C["bg"])

        self.class_score_path = tk.StringVar()
        self.exam_score_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.status_text = tk.StringVar(value="就绪")
        self.summary_text = tk.StringVar(value="选择或自动识别文件后即可开始")
        self.class_file_status = tk.StringVar(value="—")
        self.exam_file_status = tk.StringVar(value="—")
        self.output_file_status = tk.StringVar(value="—")

        self._apply_style()
        self._build_ui()
        self._auto_fill_defaults()

    # ────────── 样式注册 ──────────
    def _apply_style(self):
        style = ttk.Style()
        try:
            style.theme_use("vista")
        except Exception:
            pass

        style.configure("Card.TFrame", background=C["card"])
        style.configure("Section.TLabelframe", background=C["card"])
        style.configure("Section.TLabelframe.Label",
                        font=ui_font(10, "bold"), foreground=C["blue"])
        style.configure("Primary.TButton",
                        font=ui_font(11, "bold"), padding=(24, 10), foreground=C["blue"])
        style.map("Primary.TButton",
                  background=[("active", C["blue_light"])])
        style.configure("Ghost.TButton",
                        font=ui_font(9), padding=(14, 7))
        style.configure("SM.TButton",
                        font=ui_font(9), padding=(12, 6))

    # ────────── UI 结构 ──────────
    def _build_ui(self):
        # 外层滚动容器
        outer = tk.Frame(self.root, bg=C["bg"])
        outer.pack(fill=tk.BOTH, expand=True)

        self.canvas = tk.Canvas(outer, bg=C["bg"], highlightthickness=0, bd=0)
        sbar = ttk.Scrollbar(outer, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=sbar.set)
        sbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.scroll_frame = tk.Frame(self.canvas, bg=C["bg"])
        self.canvas_window = self.canvas.create_window((0, 0), window=self.scroll_frame, anchor="nw")
        self.scroll_frame.bind("<Configure>", self._on_scroll_frame_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self._bind_mousewheel(self.canvas)
        self._bind_mousewheel(self.scroll_frame)

        # 全局滚轮绑定（窗口内任意位置可滚动）
        self.root.bind_all("<MouseWheel>", self._on_mousewheel, add="+")
        self.root.bind_all("<Button-4>", self._on_mousewheel_linux, add="+")
        self.root.bind_all("<Button-5>", self._on_mousewheel_linux, add="+")

        # 内容区
        body = tk.Frame(self.scroll_frame, bg=C["bg"])
        body.pack(fill=tk.BOTH, expand=True, padx=20, pady=(16, 20))

        # ── 1. 顶栏 Banner ──
        self._build_banner(body)

        # ── 2. 状态卡片 ──
        self._build_status_cards(body)

        # ── 3. 主工作区：左(文件+操作) + 右(说明) ──
        self._build_workarea(body)

        # ── 4. 运行日志 ──
        self._build_log(body)

        # ── 5. 底部 ──
        self._build_footer(body)

        # 日志欢迎
        self._log("欢迎使用 v3.1 发布版 · 界面更精致，适合直接打包分发 🌸")

    # ─── Banner ───
    def _build_banner(self, parent):
        banner = tk.Frame(parent, bg=C["banner"], bd=0, highlightthickness=1,
                          highlightbackground=C["banner_deep"])
        banner.pack(fill=tk.X, pady=(0, 16))
        banner.config(padx=24, pady=18)

        title_col = tk.Frame(banner, bg=C["banner"])
        title_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tk.Label(title_col, text="学习通成绩汇总工具",
                 font=ui_font(18, "bold"), bg=C["banner"], fg=C["text_inverse"]).pack(anchor="w")
        tk.Label(title_col, text=f"{APP_SUBTITLE}  ·  {APP_VERSION}",
                 font=ui_font(9), bg=C["banner"], fg="#D3E4F7").pack(anchor="w", pady=(3, 0))
        tk.Label(title_col,
                 text="离线处理｜自动识别常见文件｜适合打包成 EXE 直接发给同学使用",
                 font=ui_font(8), bg=C["banner"], fg="#E7F1FB").pack(anchor="w", pady=(6, 0))

        pill_row = tk.Frame(title_col, bg=C["banner"])
        pill_row.pack(anchor="w", pady=(10, 0))
        for txt in ("离线可用", "自动识别", "正式样式输出"):
            tk.Label(pill_row, text=f"  {txt}  ", font=ui_font(8, "bold"),
                     bg=C["banner_deep"], fg=C["text_inverse"], padx=4, pady=3).pack(side=tk.LEFT, padx=(0, 8))

        btn_row = tk.Frame(banner, bg=C["banner"])
        btn_row.pack(side=tk.RIGHT)
        ttk.Button(btn_row, text="📂 打开默认文件夹", style="Ghost.TButton",
                   command=self._open_default_folder).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_row, text="💡 使用说明", style="Ghost.TButton",
                   command=self._show_about).pack(side=tk.LEFT)

    # ─── 状态卡片 ───
    def _build_status_cards(self, parent):
        cards_row = tk.Frame(parent, bg=C["bg"])
        cards_row.pack(fill=tk.X, pady=(0, 14))

        card_data = [
            ("class",  "📋  班级成绩表", self.class_file_status, "自动识别或手动选择", C["blue"]),
            ("exam",   "📝  期末考试表", self.exam_file_status,  "支持非标准表头定位", C["blue"]),
            ("output", "💾  输出文件",   self.output_file_status, "生成后可直接打开", C["purple"]),
        ]
        self.status_cards = {}
        self.card_elements = {}

        for i, (key, title, var, note, accent) in enumerate(card_data):
            card = tk.Frame(cards_row, bg=C["card"], bd=1, relief=tk.FLAT, highlightthickness=1,
                            highlightbackground=C["border"])
            card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(4, 4) if i == 1 else (0, 0))
            card.config(padx=14, pady=10)
            card._base_bg = C["card"]

            tk.Label(card, text=title, font=ui_font(10, "bold"), bg=C["card"], fg=C["text_primary"]).pack(anchor="w")
            val_lbl = tk.Label(card, textvariable=var, font=ui_font(12, "bold"), bg=C["card"], fg=C["text_primary"])
            val_lbl.pack(anchor="w", pady=(6, 2))
            tk.Label(card, text=note, font=ui_font(8), bg=C["card"], fg=C["text_muted"]).pack(anchor="w")

            self.status_cards[key] = card
            self.card_elements[key] = {"accent": accent}

    # ─── 主工作区 ───
    def _build_workarea(self, parent):
        work = tk.Frame(parent, bg=C["bg"])
        work.pack(fill=tk.BOTH, expand=True, pady=(0, 14))

        # 左侧面板
        left = tk.Frame(work, bg=C["bg"])
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # ① 源文件
        self._section(left, "选择文件源")
        self._build_section_frame(left)
        self._file_row(left, "📋 班级成绩表", self.class_score_path, self._pick_class_file, self.class_file_status)
        self._file_row(left, "📝 期末考试成绩表", self.exam_score_path, self._pick_exam_file, self.exam_file_status)

        # ② 输出
        self._file_row(left, "💾 汇总输出路径", self.output_path, self._pick_output, self.output_file_status, save_mode=True)

        tip_box = tk.Frame(left, bg=C["card_alt"], bd=1, relief=tk.FLAT,
                           highlightthickness=1, highlightbackground=C["border"])
        tip_box.pack(fill=tk.X, pady=(12, 14))
        tip_box.config(padx=12, pady=10)
        tk.Label(tip_box,
            text="📌 默认目录：桌面/表格输入输出/表格导入导出",
            font=ui_font(8, "bold"), bg=C["card_alt"], fg=C["text_primary"], justify="left").pack(anchor="w")
        tk.Label(tip_box,
            text="如文件名不同，可手动点击浏览选择；输出文件默认命名为《汇总成绩.xlsx》。",
            font=ui_font(8), bg=C["card_alt"], fg=C["text_muted"], justify="left").pack(anchor="w", pady=(4, 0))

        # ─ 主操作按钮 ─
        self._build_action_bar(left)

        # 右侧说明面板
        right = tk.Frame(work, bg=C["bg"], width=300)
        right.pack(side=tk.RIGHT, fill=tk.Y)
        right.grid_propagate(False)
        self._build_guide(right)

    def _section(self, parent, title):
        sep = tk.Frame(parent, bg=C["bg"])
        sep.pack(fill=tk.X, pady=(8, 6))
        tk.Label(sep, text=title, font=ui_font(11, "bold"), bg=C["bg"], fg=C["text_primary"]).pack(anchor="w")
        tk.Label(sep, text="─" * 48, fg=C["border_light"], bg=C["bg"], font=("Consolas", 7)).pack(anchor="w")

    def _build_section_frame(self, parent):
        pass  # 占位

    def _file_row(self, parent, label, var, browse_cmd, status_var, save_mode=False):
        wrapper = tk.Frame(parent, bg=C["bg"])
        wrapper.pack(fill=tk.X, pady=(0, 10))

        row = tk.Frame(wrapper, bg=C["bg"])
        row.pack(fill=tk.X)

        ttk.Label(row, text=label, width=22,
                  font=ui_font(10, "bold"), foreground=C["text_primary"],
                  background=C["bg"]).pack(side=tk.LEFT, padx=(0, 4))

        entry = tk.Entry(row, textvariable=var, font=ui_font(10),
                         bg="#FFFFFF", fg=C["text_primary"],
                         insertbackground=C["blue"], relief=tk.SOLID, bd=1)
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(4, 6), ipady=5)

        txt = "另存为…" if save_mode else "浏览…"
        ttk.Button(row, text=txt, style="SM.TButton", command=browse_cmd).pack(side=tk.LEFT, padx=(0, 4))

        # 右侧状态标签
        st_lbl = tk.Label(row, textvariable=status_var, font=ui_font(9, "bold"),
                          bg=C["bg"], fg=C["text_muted"], width=6, anchor="e")
        st_lbl.pack(side=tk.LEFT, padx=(0, 4))
        wrapper._status_label = st_lbl

    def _build_action_bar(self, parent):
        bar = tk.Frame(parent, bg=C["action_bg"], bd=1, relief=tk.FLAT, highlightthickness=1,
                       highlightbackground=C["blue_border"])
        bar.pack(fill=tk.X, pady=(4, 0))
        bar.config(padx=18, pady=14)

        txt_col = tk.Frame(bar, bg=C["action_bg"])
        txt_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tk.Label(txt_col, text="⚡ 执行操作", font=ui_font(11, "bold"), bg=C["action_bg"], fg=C["blue"]).pack(anchor="w")
        tk.Label(txt_col, text="确认文件无误后，点击主按钮即可生成发布可用的汇总成绩表",
                 font=ui_font(8), bg=C["action_bg"], fg=C["text_secondary"]).pack(anchor="w", pady=(3, 0))

        btn_row = tk.Frame(bar, bg=C["action_bg"])
        btn_row.pack(side=tk.RIGHT)

        self.merge_btn = ttk.Button(btn_row, text="▶  一键生成汇总", style="Primary.TButton",
                                    command=self._start_merge, width=16)
        self.merge_btn.pack(side=tk.LEFT, padx=(0, 8))

        self.detect_btn = ttk.Button(btn_row, text="🔄  重新识别", style="SM.TButton",
                                     command=self._auto_fill_defaults, width=12)
        self.detect_btn.pack(side=tk.LEFT, padx=(0, 6))

    def _build_guide(self, parent):
        card = tk.Frame(parent, bg=C["card"], bd=1, relief=tk.FLAT, highlightthickness=1,
                        highlightbackground=C["border"])
        card.pack(fill=tk.BOTH, expand=True)
        card.config(padx=16, pady=14)

        tk.Label(card, text="📖 使用说明", font=ui_font(11, "bold"), bg=C["card"], fg=C["text_primary"]).pack(anchor="w")
        tk.Label(card, text="─" * 32, fg=C["border_light"], bg=C["card"], font=("Consolas", 7)).pack(anchor="w", pady=(4, 10))

        self._step(card, "1", "准备文件", "将班级成绩表和期末考试表放入默认目录，或手动浏览选择")
        self._step(card, "2", "检查识别", "确认上方状态卡片显示已识别 / 已设置")
        self._step(card, "3", "一键生成", "点击主按钮，等待日志提示生成成功")
        self._step(card, "4", "查看结果", "可直接打开输出文件夹，或双击打开结果文件")

        # 特性
        tk.Label(card, text="─" * 32, fg=C["border_light"], bg=C["card"], font=("Consolas", 7)).pack(anchor="w", pady=(10, 10))
        tk.Label(card, text="✨ 工具特性", font=ui_font(10, "bold"), bg=C["card"], fg=C["blue"]).pack(anchor="w")

        features = [
            "✓  自动识别常见文件名",
            "✓  自动去除 线下录入 / 最终成绩",
            "✓  支持非标准表头定位",
            "✓  保留正式汇总表样式",
            "✓  全程离线，无需联网",
            "✓  适合打包 EXE 后直接分发"
        ]
        for f in features:
            tk.Label(card, text=f, font=ui_font(9), bg=C["card"], fg=C["text_secondary"],
                     justify="left").pack(anchor="w", pady=(3, 0))

    def _step(self, parent, num, title, desc):
        row = tk.Frame(parent, bg=C["card"])
        row.pack(fill=tk.X, pady=(0, 8))

        badge = tk.Label(row, text=num, font=ui_font(9, "bold"), bg=C["blue"], fg="#FFFFFF",
                         width=2, relief=tk.FLAT)
        badge.pack(side=tk.LEFT, padx=(0, 10), ipady=3)

        col = tk.Frame(row, bg=C["card"])
        col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tk.Label(col, text=title, font=ui_font(9, "bold"), bg=C["card"], fg=C["text_primary"]).pack(anchor="w")
        tk.Label(col, text=desc, font=ui_font(8), bg=C["card"], fg=C["text_secondary"],
                 wraplength=220, justify="left").pack(anchor="w", pady=(1, 0))

    def _build_log(self, parent):
        frame = tk.Frame(parent, bg=C["card"], bd=1, relief=tk.FLAT, highlightthickness=1,
                         highlightbackground=C["border"])
        frame.pack(fill=tk.BOTH, expand=True, pady=(0, 0))
        frame.config(padx=14, pady=10)

        hdr = tk.Frame(frame, bg=C["card"])
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text="⚙ 运行日志", font=ui_font(10, "bold"), bg=C["card"], fg=C["text_primary"]).pack(side=tk.LEFT)
        ttk.Button(hdr, text="清空", style="SM.TButton", command=self._clear_log).pack(side=tk.RIGHT)

        self.log_text = scrolledtext.ScrolledText(
            frame, height=13, font=mono_font(8),
            bg=C["term_bg"], fg="#8FA4BF",
            insertbackground="#FFFFFF",
            wrap=tk.WORD, relief=tk.FLAT,
            padx=12, pady=8, bd=0
        )
        self.log_text.pack(fill=tk.BOTH, expand=True, pady=(8, 0))
        self._bind_mousewheel(self.log_text)

        # 底部状态 & 摘要
        bottom = tk.Frame(frame, bg=C["card"])
        bottom.pack(fill=tk.X, pady=(10, 0))

        st_bar = tk.Frame(bottom, bg=C["blue_light"], padx=12, pady=8, bd=1, relief=tk.FLAT,
                          highlightthickness=1, highlightbackground=C["blue_border"])
        st_bar.pack(fill=tk.X, pady=(0, 6))
        tk.Label(st_bar, text="状态：", font=ui_font(9, "bold"), bg=C["blue_light"], fg=C["blue"]).pack(side=tk.LEFT)
        self.status_value_label = tk.Label(st_bar, textvariable=self.status_text,
                                            font=ui_font(9, "bold"), bg=C["blue_light"], fg=C["blue"])
        self.status_value_label.pack(side=tk.LEFT, padx=(4, 0))

        sm_bar = tk.Frame(bottom, bg=C["card_alt"], padx=12, pady=8, bd=1, relief=tk.FLAT,
                          highlightthickness=1, highlightbackground=C["border"])
        sm_bar.pack(fill=tk.X)
        tk.Label(sm_bar, text="摘要：", font=ui_font(9, "bold"), bg=C["card_alt"], fg=C["text_primary"]).pack(side=tk.LEFT)
        tk.Label(sm_bar, textvariable=self.summary_text, font=ui_font(9), bg=C["card_alt"],
                 fg=C["text_secondary"], wraplength=900, justify="left").pack(side=tk.LEFT, padx=(4, 0), fill=tk.X, expand=True)

        # 结果按钮
        rb = tk.Frame(bottom, bg=C["card"])
        rb.pack(fill=tk.X, pady=(8, 0))
        self.open_output_folder_btn = ttk.Button(rb, text="📂 打开输出文件夹", style="SM.TButton",
                                                  command=self._open_output_folder, width=16)
        self.open_output_folder_btn.pack(side=tk.LEFT, padx=(0, 6))
        self.open_output_file_btn = ttk.Button(rb, text="📄 打开结果文件", style="SM.TButton",
                                                command=self._open_output_file, width=14)
        self.open_output_file_btn.pack(side=tk.LEFT)

    def _build_footer(self, parent):
        ft = tk.Label(parent,
                      text=f"{APP_TITLE}  {APP_VERSION}  ·  离线运行 · 本地处理 · 默认输出到桌面 · 可打包 EXE 分发",
                      font=ui_font(8), bg=C["bg"], fg=C["text_muted"], justify="center")
        ft.pack(anchor="center", pady=(14, 0))

    # ============================================================
    #  滚动处理（不变）
    # ============================================================
    def _on_scroll_frame_configure(self, event=None):
        if hasattr(self, "canvas"):
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        if hasattr(self, "canvas") and hasattr(self, "canvas_window"):
            self.canvas.itemconfigure(self.canvas_window, width=event.width)

    def _bind_mousewheel(self, widget):
        """绑定滚轮事件到指定部件（仅用于局部精细控制）"""
        widget.bind("<MouseWheel>", self._on_mousewheel, add="+")
        widget.bind("<Button-4>", self._on_mousewheel_linux, add="+")
        widget.bind("<Button-5>", self._on_mousewheel_linux, add="+")

    def _get_mouse_widget(self):
        """获取鼠标当前所在的小部件"""
        try:
            px = self.root.winfo_pointerx()
            py = self.root.winfo_pointery()
            return self.root.winfo_containing(px, py)
        except Exception:
            return None

    def _is_inside_log(self, widget):
        """检查小部件是否在日志区域内"""
        if widget is None or not hasattr(self, 'log_text'):
            return False
        log_widget = self.log_text
        w = widget
        while w is not None:
            if w is log_widget:
                return True
            w = w.master
        return False

    def _is_log_at_top(self):
        """检查日志是否已滚动到顶部"""
        try:
            return float(self.log_text.yview()[0]) <= 0.001
        except Exception:
            return True

    def _is_log_at_bottom(self):
        """检查日志是否已滚动到底部"""
        try:
            return float(self.log_text.yview()[1]) >= 0.999
        except Exception:
            return True

    def _on_mousewheel(self, event):
        mouse_widget = self._get_mouse_widget()
        if mouse_widget is None:
            return

        # 计算滚动步数
        steps = int(-event.delta / 40)
        if steps == 0:
            steps = -3 if event.delta > 0 else 3

        # 如果在日志区，先滚日志，滚完再滚主页面
        if self._is_inside_log(mouse_widget):
            # 向下滚动
            if steps > 0:
                if self._is_log_at_bottom():
                    self.canvas.yview_scroll(steps, "units")
                else:
                    self.log_text.yview_scroll(steps, "units")
            # 向上滚动
            else:
                if self._is_log_at_top():
                    self.canvas.yview_scroll(steps, "units")
                else:
                    self.log_text.yview_scroll(steps, "units")
        else:
            self.canvas.yview_scroll(steps, "units")
        return "break"

    def _on_mousewheel_linux(self, event):
        mouse_widget = self._get_mouse_widget()
        if mouse_widget is None:
            return

        step = -3 if event.num == 4 else 3

        # 如果在日志区，先滚日志，滚完再滚主页面
        if self._is_inside_log(mouse_widget):
            if event.num == 5:  # 向下
                if self._is_log_at_bottom():
                    self.canvas.yview_scroll(step, "units")
                else:
                    self.log_text.yview_scroll(step, "units")
            else:  # 向上
                if self._is_log_at_top():
                    self.canvas.yview_scroll(step, "units")
                else:
                    self.log_text.yview_scroll(step, "units")
        else:
            self.canvas.yview_scroll(step, "units")
        return "break"

    # ============================================================
    #  状态卡片颜色更新
    # ============================================================
    def _set_card_accent(self, key, mode="default"):
        card = self.status_cards.get(key)
        if not card:
            return
        palettes = {
            "default": {"bg": C["card"], "border": C["border"], "fg": C["text_primary"]},
            "success": {"bg": C["green_bg"], "border": C["green_border"], "fg": C["green_text"]},
            "warn":    {"bg": C["orange_bg"], "border": C["orange_border"], "fg": C["orange_text"]},
            "danger":  {"bg": C["red_bg"], "border": C["red_border"], "fg": C["red_text"]},
        }
        t = palettes.get(mode, palettes["default"])
        card.config(bg=t["bg"], highlightbackground=t["border"], highlightcolor=t["border"], highlightthickness=1)
        for child in card.winfo_children():
            if isinstance(child, tk.Label):
                child.config(bg=t["bg"])

    def _refresh_action_buttons(self):
        cls = Path(self.class_score_path.get()).exists() if self.class_score_path.get().strip() else False
        exm = Path(self.exam_score_path.get()).exists() if self.exam_score_path.get().strip() else False
        out_ready = bool(self.output_path.get().strip())
        out_exists = Path(self.output_path.get()).exists() if self.output_path.get().strip() else False

        m_st = tk.NORMAL if (cls and exm and out_ready) else tk.DISABLED
        o_st = tk.NORMAL if out_exists else tk.DISABLED

        if hasattr(self, "merge_btn"):
            self.merge_btn.config(state=m_st)
        if hasattr(self, "open_output_folder_btn"):
            self.open_output_folder_btn.config(state=tk.NORMAL if out_ready else tk.DISABLED)
        if hasattr(self, "open_output_file_btn"):
            self.open_output_file_btn.config(state=o_st)

    def _update_file_status(self):
        cls = Path(self.class_score_path.get()).exists() if self.class_score_path.get().strip() else False
        exm = Path(self.exam_score_path.get()).exists() if self.exam_score_path.get().strip() else False
        out_ready = bool(self.output_path.get().strip())
        out_exists = Path(self.output_path.get()).exists() if self.output_path.get().strip() else False

        self.class_file_status.set("✓ 已识别" if cls else "—")
        self.exam_file_status.set("✓ 已识别" if exm else "—")
        self.output_file_status.set("✓ 已生成" if out_exists else ("✓ 已设置" if out_ready else "—"))

        self._set_card_accent("class", "success" if cls else "warn")
        self._set_card_accent("exam", "success" if exm else "warn")
        self._set_card_accent("output", "success" if out_exists else ("default" if out_ready else "warn"))
        self._refresh_action_buttons()

    # ============================================================
    #  文件操作
    # ============================================================
    def _find_existing_file(self, folder: Path, patterns, keyword_groups=None):
        if not folder.exists():
            return None
        for name in patterns:
            p = folder / name
            if p.exists():
                return p
        excel_files = sorted([p for p in folder.glob("*.xls*") if p.is_file()], key=lambda p: p.name)
        if keyword_groups:
            for keywords in keyword_groups:
                for p in excel_files:
                    if all(k in p.name for k in keywords):
                        return p
        return None

    def _auto_fill_defaults(self):
        folder = DEFAULT_FOLDER
        class_file = None
        exam_file = None

        if folder.exists():
            class_file = self._find_existing_file(folder, CLASS_PATTERNS, CLASS_KEYWORDS)
            exam_file = self._find_existing_file(folder, EXAM_PATTERNS, EXAM_KEYWORDS)

        if class_file:
            self.class_score_path.set(str(class_file))
        if exam_file:
            self.exam_score_path.set(str(exam_file))
        self.output_path.set(str(folder / OUTPUT_NAME))

        self._update_file_status()
        self._set_summary("已完成自动识别，可以直接开始生成")
        self._log("已执行自动识别。")
        self._log(f"默认文件夹: {folder}")
        self._log(f"班级成绩表: {self.class_score_path.get() or '未识别到'}")
        self._log(f"期末考试表: {self.exam_score_path.get() or '未识别到'}")
        self._log(f"输出文件: {self.output_path.get()}")
        self._set_status("已自动识别默认路径", C["blue"])

    def _pick_class_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")], title="选择班级成绩表")
        if path:
            self.class_score_path.set(path)
            self._update_file_status()
            self._set_summary("班级成绩表已选择")
            self._log(f"已选择班级成绩表: {path}")

    def _pick_exam_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")], title="选择期末考试成绩表")
        if path:
            self.exam_score_path.set(path)
            self._update_file_status()
            self._set_summary("期末考试成绩表已选择")
            self._log(f"已选择期末考试成绩表: {path}")

    def _pick_output(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=OUTPUT_NAME,
            title="选择汇总成绩保存位置",
        )
        if path:
            self.output_path.set(path)
            self._update_file_status()
            self._set_summary("输出位置已设置")
            self._log(f"输出文件设置为: {path}")

    def _open_default_folder(self):
        folder = DEFAULT_FOLDER
        if not folder.exists():
            messagebox.showwarning("提示", f"默认文件夹不存在：\n{folder}")
            return
        try:
            os.startfile(folder)
        except Exception as e:
            messagebox.showerror("错误", f"打开文件夹失败：\n{e}")

    def _open_output_folder(self):
        out_path = self.output_path.get().strip()
        if not out_path:
            messagebox.showwarning("提示", "请先设置输出路径。")
            return
        folder = Path(out_path).parent
        if not folder.exists():
            messagebox.showwarning("提示", f"输出文件夹不存在：\n{folder}")
            return
        try:
            os.startfile(folder)
        except Exception as e:
            messagebox.showerror("错误", f"打开输出文件夹失败：\n{e}")

    def _open_output_file(self):
        out_path = self.output_path.get().strip()
        if not out_path:
            messagebox.showwarning("提示", "请先设置输出文件路径。")
            return
        out_file = Path(out_path)
        if not out_file.exists():
            messagebox.showwarning("提示", f"结果文件不存在：\n{out_file}")
            return
        try:
            os.startfile(out_file)
        except Exception as e:
            messagebox.showerror("错误", f"打开结果文件失败：\n{e}")

    # ============================================================
    #  核心合并逻辑（与原代码一致）
    # ============================================================
    def _start_merge(self):
        class_path = self.class_score_path.get().strip()
        exam_path = self.exam_score_path.get().strip()
        out_path = self.output_path.get().strip()

        if not all([class_path, exam_path, out_path]):
            messagebox.showerror("错误", "请先确认班级成绩表、期末考试成绩表和输出路径。")
            return
        if not Path(class_path).exists():
            messagebox.showerror("错误", f"班级成绩表不存在：\n{class_path}")
            return
        if not Path(exam_path).exists():
            messagebox.showerror("错误", f"期末考试成绩表不存在：\n{exam_path}")
            return

        self.merge_btn.config(state=tk.DISABLED)
        self.detect_btn.config(state=tk.DISABLED)
        self.open_output_folder_btn.config(state=tk.DISABLED)
        self.open_output_file_btn.config(state=tk.DISABLED)

        self.log_text.delete(1.0, tk.END)
        self._log("开始生成汇总成绩...")
        self._set_status("正在处理，请稍候...", C["orange_text"])
        self._set_summary("正在分析表格并生成汇总文件")
        self._set_card_accent("output", "warn")
        threading.Thread(target=self._do_merge, args=(class_path, exam_path, out_path), daemon=True).start()

    def _do_merge(self, class_path, exam_path, out_path):
        try:
            self._log(f"班级成绩表: {class_path}")
            self._log(f"期末考试表: {exam_path}")
            self._log(f"输出文件: {out_path}")

            df1 = pd.read_excel(class_path, header=1)
            df1 = df1.dropna(subset=[df1.columns[0]], how="all").reset_index(drop=True)
            self._log(f"已读取班级成绩表：{len(df1)} 行，{len(df1.columns)} 列", "OK")

            id_col = self._find_col(df1, ["学号/工号", "学号", "工号"])
            if not id_col:
                raise ValueError("班级成绩表中找不到学号/工号列。")
            self._log(f"识别学号列：{id_col}")

            fields = {
                "章节任务点": ["章节任务点(20%)", "章节任务点"],
                "作业": ["作业(15%)", "作业"],
                "签到": ["签到(50%)", "签到"],
                "课程积分": ["课程积分(15%)", "课程积分"],
                "综合成绩": ["综合成绩", "平时综合成绩"],
            }
            found_cols = {}
            for name, candidates in fields.items():
                col = self._find_col(df1, candidates)
                found_cols[name] = col
                self._log(f"{name}: {col if col else '未找到'}", "OK" if col else "WARN")

            df2_raw = pd.read_excel(exam_path, header=None)
            self._log(f"已读取期末原始表：{df2_raw.shape[0]} 行，{df2_raw.shape[1]} 列", "OK")

            header_row = None
            for r in range(df2_raw.shape[0]):
                row_vals = df2_raw.iloc[r].astype(str).tolist()
                if any("学号/工号" in str(v) for v in row_vals):
                    header_row = r
                    break
            if header_row is None:
                raise ValueError("期末考试成绩表中找不到包含学号/工号的表头行。")
            self._log(f"识别到期末表头行为第 {header_row + 1} 行")

            df2 = pd.read_excel(exam_path, header=header_row)
            df2 = df2.dropna(subset=[df2.columns[0]], how="all").reset_index(drop=True)
            self._log(f"已读取期末正式数据：{len(df2)} 行")

            id_col2 = self._find_col(df2, ["学号/工号", "学号", "工号"])
            if not id_col2:
                raise ValueError("期末考试成绩表中找不到学号/工号列。")

            total_col = self._find_col(df2, ["总分", "客观题总分", "期末总成绩"])
            if not total_col:
                raise ValueError("期末考试成绩表中找不到总分列。")
            self._log(f"识别到期末总分列：{total_col}")

            df1[id_col] = df1[id_col].astype(str).str.replace(".0", "", regex=False).str.strip()
            df2[id_col2] = df2[id_col2].astype(str).str.replace(".0", "", regex=False).str.strip()

            score_map = df2[[id_col2, total_col]].copy()
            score_map.columns = ["学号/工号", "线上客观题成绩"]

            extract = {"学号/工号": df1[id_col]}
            rename_map = {
                "章节任务点": "章节任务点(20%)",
                "作业": "作业(15%)",
                "签到": "签到(50%)",
                "课程积分": "课程积分(15%)",
                "综合成绩": "平时综合成绩",
            }
            for name, col in found_cols.items():
                if col:
                    extract[rename_map[name]] = df1[col]

            base_df = pd.DataFrame(extract)
            result = base_df.merge(score_map, on="学号/工号", how="left")
            result.insert(0, "序号", range(1, len(result) + 1))

            drop_cols = [c for c in ["线下录入", "最终成绩"] if c in result.columns]
            if drop_cols:
                result = result.drop(columns=drop_cols)

            self._log(f"合并完成：{len(result)} 行，{len(result.columns)} 列", "OK")
            self._log(f"输出列：{list(result.columns)}")

            self._write_excel(result, out_path)
            self._log(f"已成功生成：{out_path}", "OK")
            self._set_status("生成成功", C["green_text"])
            self._set_summary(f"生成完成：共 {len(result)} 条记录，输出 {len(result.columns)} 列，可直接打开结果文件")
            self._update_file_status()
            self._set_card_accent("output", "success")
            messagebox.showinfo(
                "生成完成",
                f"汇总成绩文件已成功生成。\n\n"
                f"记录数：{len(result)} 条\n"
                f"输出列数：{len(result.columns)} 列\n\n"
                f"输出文件：\n{out_path}\n\n"
                "现在可以直接打开结果文件，或进入输出文件夹继续处理。"
            )
        except Exception as e:
            self._log(f"处理失败：{e}", "ERROR")
            self._log(traceback.format_exc(), "ERROR")
            self._set_status("处理失败", C["red_text"])
            self._set_summary("生成失败，请查看下方运行日志并检查源文件格式")
            self._set_card_accent("output", "danger")
            messagebox.showerror("错误", f"处理失败：\n\n{e}")
        finally:
            self.detect_btn.config(state=tk.NORMAL)
            self._update_file_status()

    def _find_col(self, df, candidates):
        cols = [str(c) for c in df.columns]
        for c in candidates:
            for col in cols:
                if c in col:
                    return df.columns[cols.index(col)]
        return None

    def _write_excel(self, result_df, out_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "汇总成绩"

        title_fill = PatternFill(start_color="DCEBFF", end_color="DCEBFF", fill_type="solid")
        header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        even_fill = PatternFill(start_color="EEF4FB", end_color="EEF4FB", fill_type="solid")
        thin = Side(style="thin", color="C7D3E0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        num_cols = len(result_df.columns)
        end_col_letter = openpyxl.utils.get_column_letter(num_cols)
        ws.merge_cells(f"A1:{end_col_letter}1")
        ws["A1"] = "平时成绩情况一览表"
        ws["A1"].font = Font(name="Microsoft YaHei", size=15, bold=True, color="1F2D3D")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].fill = title_fill
        ws.row_dimensions[1].height = 28

        headers = list(result_df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[2].height = 24

        for row_idx, (_, row) in enumerate(result_df.iterrows(), start=3):
            fill = odd_fill if (row_idx % 2 == 1) else even_fill
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name="Microsoft YaHei", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
                cell.fill = fill
                if col_idx == 2:
                    cell.number_format = "@"

        width_map = {
            "序号": 8,
            "学号/工号": 18,
            "章节任务点(20%)": 16,
            "作业(15%)": 14,
            "签到(50%)": 14,
            "课程积分(15%)": 16,
            "平时综合成绩": 16,
            "线上客观题成绩": 16,
        }
        for idx, col_name in enumerate(headers, start=1):
            letter = openpyxl.utils.get_column_letter(idx)
            ws.column_dimensions[letter].width = width_map.get(col_name, 16)

        ws.freeze_panes = "A3"
        ws.sheet_view.showGridLines = False
        ws.auto_filter.ref = f"A2:{end_col_letter}{len(result_df) + 2}"

        out_file = Path(out_path)
        out_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_file)

    def _log(self, msg, level="INFO"):
        prefix_map = {"INFO": "[INFO]", "OK": "[OK] ✓", "WARN": "[WARN]", "ERROR": "[ERROR]"}
        line = f"{prefix_map.get(level, '[INFO]')} {msg}"
        self.log_text.insert(tk.END, line + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def _clear_log(self):
        self.log_text.delete(1.0, tk.END)
        self._log("日志已清空。")

    def _show_about(self):
        messagebox.showinfo(
            "关于",
            f"{APP_TITLE}\n{APP_VERSION}\n\n"
            f"定位：{APP_SUBTITLE}\n"
            "特点：离线运行、自动识别、正式汇总样式输出、适合 EXE 分发\n\n"
            f"默认目录：\n{DEFAULT_FOLDER}\n\n"
            "适用场景：将学习通班级成绩表与期末考试成绩表合并为正式汇总成绩文件。\n\n"
            "发布说明：可使用 PyInstaller 打包为单文件 EXE，在 Windows 10/11 64 位环境中直接双击使用。"
        )

    def _set_status(self, msg, color=C["text_secondary"]):
        self.status_text.set(msg)
        if hasattr(self, "status_value_label"):
            self.status_value_label.config(fg=color)
        self.root.update_idletasks()

    def _set_summary(self, msg):
        self.summary_text.set(msg)
        self.root.update_idletasks()


def main():
    enable_high_dpi()
    root = tk.Tk()
    ui_scale = detect_ui_scale()
    try:
        root.tk.call("tk", "scaling", ui_scale)
    except Exception:
        pass
    ScoreMergerGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()